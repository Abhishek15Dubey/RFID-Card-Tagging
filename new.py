import keyboard
from openpyxl import load_workbook
from time import sleep, time

def convert_to_mfrc522_format(decimal_uid):
    # Convert decimal to hex string WITHOUT forcing 8 characters
    hex_string = format(decimal_uid, 'X')  # No zero padding
    
    # Reverse the byte order (little-endian to big-endian)
    # Handle odd-length hex strings by working with what we have
    if len(hex_string) % 2 != 0:
        # If odd length, process as pairs starting from the end
        bytes_list = []
        i = len(hex_string) - 2
        while i >= 0:
            bytes_list.append(hex_string[i:i+2])
            i -= 2
        # Handle the leftover single character if any
        if i == -1:
            bytes_list.append(hex_string[0])
        mfrc522_hex = ''.join(bytes_list)
    else:
        # Even length - normal byte reversal
        bytes_list = [hex_string[i:i+2] for i in range(0, len(hex_string), 2)]
        bytes_list.reverse()
        mfrc522_hex = ''.join(bytes_list)
    
    return mfrc522_hex

# === CONFIGURATION ===
file_path = r"C:\Users\ad902\Desktop\MAIN READER LIBRARY\final.xlsx3660016191" \

rfid_column = 16
start_row = 1

# === SETUP EXCEL ===
wb = load_workbook(file_path)
sheet = wb.active

# === LOAD EXISTING RFIDs ===
existing_rfids = set()
row = start_row
while sheet.cell(row=row, column=rfid_column).value is not None:
    existing_rfids.add(str(sheet.cell(row=row, column=rfid_column).value))
    row += 1

current_row = row

# === CAPTURE VARIABLES ===
captured_digits = []
last_key_time = 0

def on_key_event(e):
    global captured_digits, last_key_time
    
    current_time = time()
    
    # If too much time passed since last key, reset (new scan)
    if current_time - last_key_time > 1.0 and captured_digits:
        process_completed_uid()
    
    # Capture number keys only
    if e.event_type == keyboard.KEY_DOWN and e.name in '0123456789':
        captured_digits.append(e.name)
        last_key_time = current_time
    
    # If Enter is pressed, process immediately
    elif e.event_type == keyboard.KEY_DOWN and e.name == 'enter':
        process_completed_uid()

def process_completed_uid():
    global captured_digits, current_row
    
    if len(captured_digits) >= 6:  # Minimum length for RFID
        decimal_str = ''.join(captured_digits)
        
        try:
            decimal_uid = int(decimal_str)
            mfrc522_hex = convert_to_mfrc522_format(decimal_uid)
            
            if mfrc522_hex in existing_rfids:
                print(f"⏭️  DUPLICATE: {decimal_str} -> {mfrc522_hex}")
            else:
                print(f"✅ Row {current_row}: {decimal_str} -> {mfrc522_hex}")
                sheet.cell(row=current_row, column=rfid_column).value = mfrc522_hex
                existing_rfids.add(mfrc522_hex)
                current_row += 1
                wb.save(file_path)
                print("💾 Saved to Excel!")
                
        except ValueError:
            print(f"❌ Invalid UID: {decimal_str}")
    
    # Reset for next scan
    captured_digits.clear()

# === MAIN PROGRAM ===
print(f"📋 Resuming from row {current_row}")
print("🔍 RFID Auto-Capture Active!")
print("💡 Scan RFID cards - I'll automatically capture and convert them")
print("⏹️  Press ESC to stop\n")

# Start listening to keyboard
keyboard.hook(on_key_event)

try:
    # Keep the program running
    while True:
        sleep(0.1)
        # Auto-process if we have digits and no new input for a while
        if captured_digits and (time() - last_key_time > 0.5):
            process_completed_uid()
            
except KeyboardInterrupt:
    print("\n🛑 Stopped by user")

finally:
    keyboard.unhook_all()
    wb.save(file_path)
    wb.close()
    print("💾 Excel file saved successfully")
    print("👋 Program ended")